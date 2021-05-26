# @Date  : 2021/3/12
# @Author: Hugh
# @Email : 609799548@qq.com

"""
封装 Excel 数据获取

使用说明
<1> 配置 Header 类中，各字段对应 Excel 表中的列号
<2> excel = Excel(excel_path)
<3> excel.data: 可获取对应 Excel 文件中的所有数据
"""

import enum

import openpyxl


@enum.unique
class Header(enum.Enum):
    CASE_TITLE = 'a'
    CASE_URL = 'b'
    CASE_METHOD = 'c'
    CASE_HEADER = 'd'
    CASE_DATA = 'e'
    CASE_EXPECT = 'f'
    CASE_SETUP = 'g'
    CASE_TEARDOWN = 'H'
    CASE_EXTRACT = 'I'
    CASE_RUN = 'J'
    CASE_TIMEOUT = 'K'


class OperationExcel:
    def __init__(self, excel_path):
        self.wb = openpyxl.load_workbook(excel_path)

    @staticmethod
    def __column_char_to_number(column_char):
        return ord(column_char.lower()) - 96

    @property
    def sheet_names(self):
        return self.wb.sheetnames

    def get_sheet_data_by_name(self, name):
        return self.__get_sheet_data(name)

    @property
    def data(self):
        return {sheet_name: self.get_sheet_data_by_name(sheet_name) for sheet_name in self.sheet_names}

    def __get_header(self, name, row=1):
        result = []
        column = 1
        while True:
            cell_value = self.wb[name].cell(row=row, column=column).value
            if not cell_value:
                break
            result.append(cell_value)
            column += 1
        return result

    def __get_sheet_data(self, name):
        sheet = self.wb[name]
        body = []
        row = 2
        while True:
            row_data = {}
            for member in Header.__members__.values():
                row_data[member] = sheet.cell(row=row, column=self.__column_char_to_number(member.value)).value
            if not any(row_data.values()):
                break
            row_data['EXCEL_ROW'] = row
            row += 1
            body.append(row_data)
        return body

    def __del__(self):
        self.wb.close()
